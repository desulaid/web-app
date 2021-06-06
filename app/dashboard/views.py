import os
from datetime import datetime

from flask import Blueprint, send_from_directory, current_app as app
from flask import render_template, request, flash, redirect, url_for, abort
from flask_login import login_required, current_user as user
from openpyxl import Workbook
from openpyxl.styles import Font

from app.database import db, Group, Teacher, Profile, Student, Post, Task, File

dashboard = Blueprint('dashboard', __name__, template_folder='templates', url_prefix='/dashboard')


@dashboard.route('/settings', methods=['GET'])
@login_required
def settings():
    groups = Group.query.all()
    teachers = Teacher.query.all()

    context = dict(
        title='Настройки профиля',
        header='Настройки',
        user=user,
        groups=groups,
        teachers=teachers
    )

    return render_template('dashboard/settings.html', **context)


@dashboard.route('/save/settings/profile/<string:user_login>', methods=['POST'])
@login_required
def save_profile_settings(user_login: str):
    profile = Profile.query.filter_by(login=user_login).first()

    error = None
    name = request.form['name']
    login = request.form['login']
    password = request.form['password']
    password_confirm = request.form['password-confirm']

    if db.session.query(
            Profile.query.filter_by(login=login).exists().where(profile.login != login)
    ).scalar():
        error = 'Логин уже занят'

    elif password != password_confirm:
        error = 'Пароли не совпадают'
    elif not len(password):
        error = 'Пароль не может быть пустым'

    if error is None:
        profile.name = name
        profile.login = login
        profile.password = password

        db.session.commit()

        flash('Настройки обновлены', 'success')
    else:
        flash(error, 'warning')

    return redirect(url_for('.settings'))


@dashboard.route('/save/settings/professional/<string:user_login>', methods=['POST'])
@login_required
def save_proff_settings(user_login: str):
    profile = Profile.query.filter_by(login=user_login).first()

    error = None
    group_id = request.form['group-id']
    teacher_id = request.form['group-teacher']
    group_name = Group.query.get(group_id)

    if db.session.query(
            Profile.query.filter_by(group=group_id,
                                    teacher=teacher_id).exists()
    ).scalar():
        error = f'Староста группы {group_name.name} уже назначен'

    if error is None:
        profile.group = group_id
        profile.teacher = teacher_id
        db.session.commit()
    else:
        flash(error, 'warning')

    return redirect(url_for('.settings'))


@dashboard.route('/verify', methods=['GET'])
@login_required
def verify():
    if user.id != 1:
        abort(404)

    profiles = Profile.query.filter_by(verify=False).all()

    context = dict(
        title='Подтверждение профилей',
        header='Верификация',
        user=user,
        profiles=profiles
    )

    return render_template('dashboard/verify.html', **context)


@dashboard.route('/verifying/<string:login>', methods=['POST'])
def verifying(login: str):
    status = True if request.form['status'] == 'yes' else False

    profile = Profile.query.filter_by(login=login).first()

    if status:
        profile.role = 1
        profile.verify = True
        db.session.commit()

        flash(f'Аккаунт {login} подтвержен', 'success')
    else:
        db.session.delete(profile)
        db.session.commit()

        flash(f'Аккаунт {login} удален', 'warning')

    return redirect(url_for('.verify'))


@dashboard.route('/group', methods=['GET'])
@login_required
def group():
    from json import dumps

    students = Student.query.filter_by(profile=user.id).all()

    context = dict(
        title='Студенты моей группы',
        header='Моя группа',
        data=dumps([i.name for i in students]),
        user=user,
    )

    return render_template('dashboard/group.html', **context)


@dashboard.route('/group/save/<int:id>', methods=['POST'])
@login_required
def group_save(id: int):
    students = Student.query.filter_by(profile=id).all()

    if not students:
        for key in request.form:
            db.session.add(Student(name=request.form[key],
                                   profile_id=id))
    else:
        form = [request.form[i] for i in request.form]
        table = [i.name for i in students]
        over_rows = list(set(form) ^ set(table))

        for item in over_rows:
            student = Student.query.filter_by(profile=id,
                                              name=item).first()
            if not student:
                db.session.add(Student(name=item,
                                       profile_id=id))
                flash(f'Добавлен студент {item}', 'success')
            else:
                flash(f'Удален студент {item}', 'danger')

                posts = Post.query.filter_by(profile=id).all()
                for post in posts:
                    tasks = post.tasks.filter_by(student=id).all()
                    for task in tasks:
                        db.session.delete(task)

                flash(f'Все связанные со студентом {item} записи удалены', 'danger')
                db.session.delete(student)

    db.session.commit()

    return redirect(url_for('.group'))


@dashboard.route('/create', methods=['GET'])
@login_required
def create_post():
    students = Student.query.filter_by(profile=user.id).all()

    context = dict(
        title='Добавить новую запись',
        header='Новая запись',
        students=students,
        user=user
    )

    return render_template('dashboard/create_post.html', **context)


@dashboard.route('/create/save/<int:id>', methods=['POST'])
@login_required
def save_post(id: int):
    from re import match
    from datetime import datetime

    form = dict(request.form)
    students = {}
    name = form['post-name']

    del form['post-name']
    if not name:
        flash('Название записи обязатнльео к заполнению', 'warning')
    elif not form:
        flash('Вы не можете создать запись буз студентов', 'warning')
    else:
        for key in form:
            data = match(r'student\-(\d+)\-(.+)', key)

            # Страшна, очень страшна
            try:
                if form[f'student-{data.group(1)}-attended'] == 'yes':
                    item = [True, form[f'student-{data.group(1)}-comment']]
            except KeyError:
                item = [False, form[f'student-{data.group(1)}-comment']]

            students[f'{data.group(1)}'] = item

        post = Post(profile=id)

        for i in students:
            post.tasks.append(Task(title=name,
                                   attended=students[i][0],
                                   comment=students[i][1],
                                   datetime=datetime.now(),
                                   student_id=int(i)))
            db.session.add(post)

        db.session.commit()

        flash('Запись добавлена', 'success')

    return redirect(url_for('.create_post'))


@dashboard.route('/users', methods=['GET'])
@login_required
def users():
    if user.id != 1:
        abort(404)

    profiles = Profile.query.filter_by(verify=True).all()

    context = dict(
        title='Зарегистрированные пользователи',
        header='Аккаунты',
        user=user,
        profiles=profiles
    )

    return render_template('dashboard/profiles.html', **context)


@dashboard.route('/users/delete/<name>', methods=['GET'])
@login_required
def users_delete(name):
    profile = Profile.query.filter_by(login=name).first()

    if not profile:
        return abort(404)

    db.session.delete(profile)
    db.session.commit()

    return redirect(url_for('.users'))


@dashboard.route('/report', methods=['POST'])
@login_required
def report():
    wb = Workbook()
    header_style = Font(bold=True)
    main_page = wb.active

    group = Group.query.get(user.group)
    time_create = datetime.now().strftime("%m_%d_%Y_%H_%M")
    doc = f'{user.login}_{time_create}.xlsx'

    date_from_str: datetime = lambda x: datetime.strptime(x, '%Y-%m-%d')

    start_date = date_from_str(request.form['start-date'])
    finish_date = date_from_str(request.form['finish-date'])

    students = Student.query.filter_by(profile=user.id).all()
    data = []

    main_page.title = f'Группа {group.name}'
    main_page['A1'] = 'Староста'
    main_page['A1'].font = header_style
    main_page['B1'] = f'{user.name}'

    teacher = Teacher.query.get(user.teacher)
    main_page['A2'] = 'Классный руководитель'
    main_page['A2'].font = header_style
    main_page['B2'] = f'{teacher.name}'

    main_page['A3'] = 'Отчет сформирован из системы мониторинга'
    main_page['A3'].font = header_style
    main_page['B3'] = datetime.now()

    for student in students:
        tasks = Task.query.filter_by(student=student.id).where(
            start_date <= Task.datetime).where(finish_date >= Task.datetime).all()

        page = wb.create_sheet(title=student.name)

        page_headers = [
            'Дата и время',
            'Название',
            'Присуствовал',
            'Причина'
        ]

        for col, val in enumerate(page_headers, start=1):
            cell = page.cell(column=col, row=1, value=val)
            cell.font = header_style

        lessons = []

        for row, task in enumerate(tasks, start=3):
            columns = [
                task.datetime,
                task.title,
                'Да' if task.attended else 'Нет',
                task.comment,
            ]

            for col, val in enumerate(columns, start=1):
                page.cell(column=col, row=row, value=val)

        data.append({
            'name': student.name,
            'lessons': lessons
        })

    if not data:
        flash('Нет данных для отчета', 'warning')
    else:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], doc)
        wb.save(file_path)

        db.session.add(File(name=doc, profile=user.id))
        db.session.commit()

    return redirect(url_for('.my_files'))


@dashboard.route('/files', methods=['GET'])
@login_required
def my_files():
    files = File.query.filter_by(profile=user.id).all()

    context = dict(
        user=user,
        title='Файлы',
        header='Мои отчеты',
        files=files
    )

    return render_template('dashboard/files.html', **context)


@dashboard.route('/download/file/<filename>', methods=['GET'])
@login_required
def download_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@dashboard.route('/delete/file/<filename>', methods=['GET'])
@login_required
def delete_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if os.path.exists(file_path):
        os.remove(file_path)

        file_db = File.query.filter_by(name=filename).first()

        if file_db:
            db.session.delete(file_db)

        db.session.commit()

        flash(f'Файл {filename} успешно удален', 'success')
    else:
        flash (f'Файл {filename} не сущесвует в базе данных', 'danger')
    return redirect(url_for('.my_files'))
